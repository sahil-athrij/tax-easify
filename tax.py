import openpyxl

def openwb(name):
    wb = openpyxl.load_workbook(name)
    return wb

def replace(wb,strow,stclm):
    sheetname = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheetname[0])

    row = sheet.max_row-3
    column = sheet.max_column

    for i in range(strow,row):
        value = sheet.cell(row=i , column=stclm).value
        newval = ''
        for j in value:
            if j in "!@#$%^&*()_+-=/|\\?<>.__":
                newval+=' '
            else:
                newval+=j
        st = 'A'+str(i-strow+1)
    
        sheet2 = wb.get_sheet_by_name(sheetname[1])
        sheet2[st]= newval
        sheet2['B'+str(i-strow+1)]= sheet.cell(row= i , column=column-1).value
        sheet2['C'+str(i-strow+1)]= sheet.cell(row= i , column=column).value



def calculateret(wb,strow,taxclm):
    
    sheetname = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheetname[0])

    row = sheet.max_row-3
    column = sheet.max_column

    taxtot =0
    taxind = [0,0,0,0]
    tax =[0,1,5,14.5]
    error =[]


    for i in range(strow,row):

        taxrate = sheet.cell(row=i , column=taxclm).value
        price = sheet.cell(row=i , column=column).value
        
        if taxrate ==0:
            taxtot+=price
            taxind[0]+=price
        
        elif taxrate ==1:
            taxtot+=price
            taxind[1]+=price

        elif taxrate ==5:
            taxtot+=price
            taxind[2]+=price
        
        
        elif taxrate ==14.5:
            taxtot+=price
            taxind[3]+=price

        else:
            error.append(i)
            taxtot+=price

        sheet2 = wb.get_sheet_by_name(sheetname[2])
        for k in range(1,5):
            sheet2[chr(65)+str(k)] = 'tax rate of '+str(tax[k-1])
            sheet2[chr(66)+str(k)] = (tax[k-1])
            sheet2[chr(67)+str(k)] = taxind[k-1]

        sheet2[chr(65)+str(k+1)] = 'total'
        sheet2[chr(67)+str(k+1)] = taxtot
        if error !=[]:
            sheet2[chr(66)+str(k+3)] = 'Errors'
            k=k+4
            for l in range(k,k+len(error)):
                for m in range(1,7):
                    val = sheet.cell(row=error[l-k] , column=m).value
                    sheet2[chr(m+64)+str(l)]=val
            
        
def save(wb,name):
    wb.save(name)
